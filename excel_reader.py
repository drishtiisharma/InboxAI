from openpyxl import load_workbook

def extract_text_from_xlsx(file_path, max_rows=20):
<<<<<<< HEAD
    wb = load_workbook(file_path, data_only=True)
    text = []

    for sheet in wb.worksheets:
        text.append(f"Sheet: {sheet.title}")

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        headers = rows[0]
        text.append("Columns: " + ", ".join(str(h) for h in headers if h))

        for row in rows[1:max_rows]:
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                text.append(" | ".join(row_text))
=======
    text = []

    try:
        wb = load_workbook(file_path, data_only=True)

        for sheet in wb.worksheets:
            text.append(f"Sheet: {sheet.title}")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() for h in rows[0] if h is not None]
            if headers:
                text.append("Columns: " + ", ".join(headers))

            text.append(f"Showing up to {max_rows} rows")

            for row in rows[1 : max_rows + 1]:
                if not any(row):
                    continue

                row_text = [str(cell).strip() for cell in row if cell is not None]
                if row_text:
                    text.append(" | ".join(row_text))

        wb.close()

    except Exception as e:
        return f"[ERROR reading Excel file: {str(e)}]"
>>>>>>> backend

    return "\n".join(text)
